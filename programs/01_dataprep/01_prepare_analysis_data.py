import sys
from pathlib import Path
import pandas as pd
import numpy as np
import datetime
import calendar
from openpyxl import load_workbook

PROGRAMS_DIR = Path(__file__).resolve().parents[1]
if str(PROGRAMS_DIR) not in sys.path:
    sys.path.insert(0, str(PROGRAMS_DIR))

from config import (
    BLOOMBERG_FILE,
    DVOL_FILE,
    MACRO_FILE,
    SAMPLE_START,
    ANALYSIS_DATA_DIR,
)

def create_event_dummy(event_dates, date_series, window=1):
    flagged_dates = set()
    for d in event_dates:
        for delta in range(-window, window + 1):
            flagged_dates.add((d + pd.Timedelta(days=delta)).date())
    return date_series.apply(lambda x: x.date()).isin(flagged_dates).astype(int)

def get_last_friday(year, month):
    cal = calendar.monthcalendar(year, month)
    fridays = [week[4] for week in cal if week[4] != 0]
    return pd.Timestamp(year=year, month=month, day=max(fridays))

def load_bloomberg_sheet(workbook, sheet_name, column_name):
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    data = [(r[0], r[1]) for r in rows if isinstance(r[0], datetime.datetime) and r[1] is not None]
    df = pd.DataFrame(data, columns=['date', column_name])
    df['date'] = pd.to_datetime(df['date']).dt.normalize()
    return df

bloomberg = load_workbook(BLOOMBERG_FILE, read_only=True)
vix_data = load_bloomberg_sheet(bloomberg, 'VIX', 'vix')
sp500_data = load_bloomberg_sheet(bloomberg, 'S&P 500', 'sp500')
bitcoin_prices = load_bloomberg_sheet(bloomberg, 'Bitcoin spot USD', 'btc_price')
dxy_data = load_bloomberg_sheet(bloomberg, 'US Dollar Index', 'dxy')
yield_10yr = load_bloomberg_sheet(bloomberg, 'US 10yr Yield', 'y10')
vvix_data = load_bloomberg_sheet(bloomberg, 'VVIX', 'vvix')
move_data = load_bloomberg_sheet(bloomberg, 'MOVE', 'move')

for price_df, price_col, return_col in [
    (sp500_data, 'sp500', 'sp500_ret'),
    (bitcoin_prices, 'btc_price', 'btc_ret'),
    (dxy_data, 'dxy', 'dxy_ret'),
]:
    price_df.sort_values('date', inplace=True)
    price_df[return_col] = np.log(price_df[price_col] / price_df[price_col].shift(1)) * 100

yield_10yr.sort_values('date', inplace=True)
yield_10yr['dy10'] = yield_10yr['y10'].diff()

dvol_raw = pd.read_csv(DVOL_FILE)
dvol_raw['date'] = pd.to_datetime(dvol_raw['date'])
dvol_combined = (
    dvol_raw[dvol_raw.currency == 'BTC'][['date', 'close']]
    .rename(columns={'close': 'btc_dvol'})
    .sort_values('date')
)

daily_data = (
    vix_data
    .merge(sp500_data[['date', 'sp500_ret']], on='date', how='left')
    .merge(dxy_data[['date', 'dxy_ret']], on='date', how='left')
    .merge(yield_10yr[['date', 'dy10']], on='date', how='left')
    .merge(vvix_data[['date', 'vvix']], on='date', how='left')
    .merge(move_data[['date', 'move']], on='date', how='left')
    .merge(bitcoin_prices[['date', 'btc_ret']], on='date', how='left')
    .merge(dvol_combined, on='date', how='inner')
    .dropna(subset=['btc_dvol', 'vix', 'sp500_ret', 'btc_ret'])
)

daily_data.sort_values('date', inplace=True)
daily_data.reset_index(drop=True, inplace=True)

daily_data = daily_data[daily_data.date >= SAMPLE_START].copy()
daily_data.reset_index(drop=True, inplace=True)

daily_data['btc_realized_vol'] = daily_data['btc_ret'].rolling(7).std()
daily_data['dvol_change'] = daily_data['btc_dvol'].diff()
daily_data['dvol_lag'] = daily_data['btc_dvol'].shift(1)
daily_data['is_friday'] = (daily_data['date'].dt.dayofweek == 4).astype(int)

def extract_event_dates(workbook, sheet_name, event_label):
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))[1:]
    return pd.to_datetime(sorted(set(
        row[0].date() for row in rows
        if row[0] and isinstance(row[0], datetime.datetime)
        and row[4] and str(row[4]).strip() == event_label
    )))

macro_workbook = load_workbook(MACRO_FILE, read_only=True)
fomc_dates = extract_event_dates(
    macro_workbook,
    'FOMC',
    'FOMC Rate Decision (Upper Bound)'
)
cpi_dates = extract_event_dates(
    macro_workbook,
    'CPI',
    'CPI MoM'
)
gdp_dates = extract_event_dates(
    macro_workbook,
    'GDP ',
    'GDP Annualized QoQ'
)

daily_data['fomc_dummy'] = create_event_dummy(
    fomc_dates,
    daily_data['date'],
    window=1
)
daily_data['cpi_dummy'] = create_event_dummy(
    cpi_dates,
    daily_data['date'],
    window=1
)
daily_data['gdp_dummy'] = create_event_dummy(
    gdp_dates,
    daily_data['date'],
    window=1
)

all_expiry_dates = []
yr, mo = daily_data.date.min().year, daily_data.date.min().month

while True:
    expiry = get_last_friday(yr, mo)

    if expiry > daily_data.date.max():
        break

    if expiry >= daily_data.date.min():
        all_expiry_dates.append({
            'expiry_date': expiry,
            'expiry_id': f'{yr}-{mo:02d}',
            'is_quarterly': int(mo in [3, 6, 9, 12]),
        })

    mo += 1

    if mo > 12:
        mo, yr = 1, yr + 1

all_expiry_dates = pd.DataFrame(all_expiry_dates)

trading_calendar = (
    daily_data['date']
    .sort_values()
    .values
    .astype('datetime64[D]')
)

event_window = 14

valid_events = []

for _, event in all_expiry_dates.iterrows():

    event_index = np.searchsorted(
        trading_calendar,
        np.datetime64(event['expiry_date'], 'D')
    )

    days_present = sum(
        1
        for offset in range(-event_window, event_window + 1)
        if 0 <= event_index + offset < len(trading_calendar)
        and len(
            daily_data[
                daily_data.date
                == pd.Timestamp(trading_calendar[event_index + offset])
            ]
        ) > 0
    )

    if days_present == 2 * event_window + 1:
        valid_events.append(event)

monthly_expiries = pd.DataFrame(valid_events)

us_holiday_expiries = [pd.Timestamp('2024-03-29')]

monthly_expiries = monthly_expiries[
    ~monthly_expiries.expiry_date.isin(us_holiday_expiries)
].reset_index(drop=True)

num_events = len(monthly_expiries)
num_quarterly = monthly_expiries.is_quarterly.sum()
num_regular = num_events - num_quarterly

fomc_window_dates = set(
    d + pd.Timedelta(days=delta)
    for d in fomc_dates
    for delta in [-1, 0, 1]
)

fomc_conflicts = monthly_expiries[
    monthly_expiries.expiry_date.isin(fomc_window_dates)
]

daily_expiries = all_expiry_dates[
    all_expiry_dates.expiry_date.isin(daily_data.date)
].copy().reset_index(drop=True)

daily_data['monthly_expiry'] = (
    daily_data['date']
    .isin(daily_expiries.expiry_date)
    .astype(int)
)

daily_data['quarterly_expiry'] = (
    daily_data['date']
    .isin(
        daily_expiries[
            daily_expiries.is_quarterly == 1
        ].expiry_date
    )
    .astype(int)
)

daily_data['regular_monthly'] = (
    (daily_data.monthly_expiry == 1)
    & (daily_data.quarterly_expiry == 0)
).astype(int)

def build_event_panel(monthly_expiries, trading_calendar, daily_data, window):
    rows = []

    for _, event in monthly_expiries.iterrows():

        event_index = np.searchsorted(
            trading_calendar,
            np.datetime64(event['expiry_date'], 'D')
        )

        for offset in range(-window, window + 1):

            idx = event_index + offset

            if idx < 0 or idx >= len(trading_calendar):
                continue

            actual_date = pd.Timestamp(trading_calendar[idx])

            row = daily_data[
                daily_data.date == actual_date
            ]

            if len(row) == 0:
                continue

            r = row.iloc[0]

            rows.append({
                'date': actual_date,
                'expiry_id': event['expiry_id'],
                'is_quarterly': event['is_quarterly'],
                'dte': offset,
                'post': int(offset >= 0),
                'dte_post': offset * int(offset >= 0),
                'btc_dvol': r['btc_dvol'],
                'dvol_change': r['dvol_change'],
                'vix': r['vix'],
                'sp500_ret': r['sp500_ret'],
                'btc_ret': r['btc_ret'],
                'dxy_ret': r.get('dxy_ret', np.nan),
                'dy10': r.get('dy10', np.nan),
                'vvix': r.get('vvix', np.nan),
                'move': r.get('move', np.nan),
                'btc_realized_vol': r.get('btc_realized_vol', np.nan),
                'fomc_dummy': r['fomc_dummy'],
                'cpi_dummy': r['cpi_dummy'],
                'gdp_dummy': r['gdp_dummy'],
            })

    panel = pd.DataFrame(rows)

    for col in panel.columns:
        if not pd.api.types.is_numeric_dtype(panel[col]):
            panel[col] = panel[col].astype(object)

    panel['quarterly_post'] = (
        panel['is_quarterly'] * panel['post']
    )

    panel['quarterly_dte_post'] = (
        panel['is_quarterly'] * panel['dte_post']
    )

    pre_mean = (
        panel[panel.dte < 0]
        .groupby('expiry_id')['btc_dvol']
        .mean()
    )

    panel['pre_event_baseline'] = (
        panel['expiry_id'].map(pre_mean)
    )

    panel['abnormal_dvol'] = (
        panel['btc_dvol']
        - panel['pre_event_baseline']
    )

    return panel

event_panel = build_event_panel(
    monthly_expiries,
    trading_calendar,
    daily_data,
    event_window
)

event_panel_wide = build_event_panel(
    monthly_expiries,
    trading_calendar,
    daily_data,
    20
)

counts_wide = (
    event_panel_wide
    .groupby('expiry_id')
    .size()
)

complete_20_ids = counts_wide[
    counts_wide == 41
].index

event_panel_wide = event_panel_wide[
    event_panel_wide.expiry_id.isin(
        complete_20_ids
    )
].copy()

ml_rows = []

for _, event in monthly_expiries.iterrows():

    event_date = event['expiry_date']

    event_index = np.searchsorted(
        trading_calendar,
        np.datetime64(event_date, 'D')
    )

    pre_expiry_rows = []
    expiry_dvol = None

    for offset in range(-event_window, 1):

        idx = event_index + offset

        if idx < 0 or idx >= len(trading_calendar):
            continue

        actual_date = pd.Timestamp(
            trading_calendar[idx]
        )

        row = daily_data[
            daily_data.date == actual_date
        ]

        if len(row) == 0:
            continue

        r = row.iloc[0]

        if offset < 0:
            pre_expiry_rows.append(
                r.to_dict()
            )
        else:
            expiry_dvol = r['btc_dvol']

    if not pre_expiry_rows or expiry_dvol is None:
        continue

    pre_expiry = (
        pd.DataFrame(pre_expiry_rows)
        .drop(columns=['date'], errors='ignore')
    )

    pre_expiry = pre_expiry.apply(
        pd.to_numeric,
        errors='coerce'
    )

    dvol_values = pre_expiry['btc_dvol'].values
    vix_values = pre_expiry['vix'].values

    ml_rows.append({
        'expiry_id': event['expiry_id'],
        'expiry_date': event_date,
        'is_quarterly': event['is_quarterly'],

        'abnormal_dvol':
            expiry_dvol
            - pre_expiry['btc_dvol'].mean(),

        'dvol_level_5d':
            pre_expiry['btc_dvol']
            .tail(5)
            .mean(),

        'dvol_change_5d':
            dvol_values[-1] - dvol_values[-6]
            if len(dvol_values) >= 6
            else np.nan,

        'dvol_change_10d':
            dvol_values[-1] - dvol_values[-11]
            if len(dvol_values) >= 11
            else np.nan,

        'dvol_std_10d':
            pre_expiry['btc_dvol']
            .tail(10)
            .std(),

        'dvol_vix_spread':
            pre_expiry['btc_dvol']
            .tail(5)
            .mean()
            - pre_expiry['vix']
            .tail(5)
            .mean(),

        'vix_level_5d':
            pre_expiry['vix']
            .tail(5)
            .mean(),

        'vix_change_5d':
            vix_values[-1] - vix_values[-6]
            if len(vix_values) >= 6
            else np.nan,

        'btc_rvol_5d':
            pre_expiry['btc_ret']
            .tail(5)
            .std(),

        'btc_ret_cum_5d':
            pre_expiry['btc_ret']
            .tail(5)
            .sum(),

        'sp500_ret_5d':
            pre_expiry['sp500_ret']
            .tail(5)
            .sum(),

        'dxy_ret_5d':
            pre_expiry['dxy_ret']
            .tail(5)
            .sum()
            if 'dxy_ret' in pre_expiry.columns
            else np.nan,

        'vvix_5d':
            pre_expiry['vvix']
            .tail(5)
            .mean()
            if 'vvix' in pre_expiry.columns
            else np.nan,

        'move_5d':
            pre_expiry['move']
            .tail(5)
            .mean()
            if 'move' in pre_expiry.columns
            else np.nan,
    })

ml_dataset = pd.DataFrame(ml_rows).dropna()

daily_data.to_csv(
    ANALYSIS_DATA_DIR / 'daily_data.csv',
    index=False
)

all_expiry_dates.to_csv(
    ANALYSIS_DATA_DIR / 'all_expiry_dates.csv',
    index=False
)

daily_expiries.to_csv(
    ANALYSIS_DATA_DIR / 'daily_expiries.csv',
    index=False
)

monthly_expiries.to_csv(
    ANALYSIS_DATA_DIR / 'monthly_expiries.csv',
    index=False
)

event_panel.to_csv(
    ANALYSIS_DATA_DIR / 'event_panel_14.csv',
    index=False
)

event_panel_wide.to_csv(
    ANALYSIS_DATA_DIR / 'event_panel_20.csv',
    index=False
)

ml_dataset.to_csv(
    ANALYSIS_DATA_DIR / 'ml_dataset.csv',
    index=False
)

pd.DataFrame({'date': fomc_dates}).to_csv(
    ANALYSIS_DATA_DIR / 'fomc_dates.csv',
    index=False
)

pd.DataFrame({'date': cpi_dates}).to_csv(
    ANALYSIS_DATA_DIR / 'cpi_dates.csv',
    index=False
)

pd.DataFrame({'date': gdp_dates}).to_csv(
    ANALYSIS_DATA_DIR / 'gdp_dates.csv',
    index=False
)

assert len(daily_data) == 1003
assert len(monthly_expiries) == 46
assert int(num_quarterly) == 14
assert int(num_regular) == 32
assert len(event_panel) == 1334
assert event_panel_wide['expiry_id'].nunique() == 45
assert len(event_panel_wide) == 1845
assert len(ml_dataset) == 46
assert len(fomc_dates) == 32
assert len(cpi_dates) == 46
assert len(gdp_dates) == 46
assert len(fomc_conflicts) == 0

print("Data preparation completed successfully.")
print(f"Daily observations: {len(daily_data):,}")
print(f"Expiration events: {num_events}")
print(f"Regular expirations: {num_regular}")
print(f"Quarterly expirations: {num_quarterly}")
print(f"14-day event panel observations: {len(event_panel):,}")
print(f"20-day event panel observations: {len(event_panel_wide):,}")
print(f"Machine learning observations: {len(ml_dataset)}")